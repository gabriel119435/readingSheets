import csv

from openpyxl import load_workbook


def parse_bool(value):
    if value is None:
        return None
    v = str(value).strip().lower()
    if v in ("true", "t", "1"):
        return True
    if v in ("false", "f", "0"):
        return False
    return None


def parse_int(value):
    try:
        return int(value)
    except (TypeError, ValueError):
        return None


def parse_stack(value):
    if not value:
        return ""
    elements = [s.strip() for s in str(value).replace(";", ",").split(",")]
    return ", ".join(sorted(set(elements)))


def normalize_row(data):
    return {
        "name": data.get("name"),
        "city": data.get("city"),
        "age": parse_int(data.get("age")),
        "gender": data.get("gender"),
        "disability": parse_bool(data.get("disability")),
        "salary_expectation": parse_int(data.get("salary_expectation")),
        "tech_stack": parse_stack(data.get("tech_stack")),
    }


def read_csv(file_obj):
    reader = csv.DictReader((line.decode("utf-8") for line in file_obj.stream))
    return [normalize_row(r) for r in reader]


def read_xlsx(file_obj):
    wb = load_workbook(file_obj, read_only=True)
    normalized = []

    for sheet in wb.worksheets:
        rows = sheet.iter_rows(values_only=True)
        headers = [str(h).strip() for h in next(rows)]

        for row in rows:
            data = dict(zip(headers, row))
            normalized.append(normalize_row(data))

    return normalized
